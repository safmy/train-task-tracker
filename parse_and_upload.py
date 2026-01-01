#!/usr/bin/env python3
"""
Parse all WorktoSheets Excel files and upload tasks to Supabase
"""

import openpyxl
import os
import re
from datetime import datetime, time as dt_time
from supabase import create_client

# Supabase configuration
SUPABASE_URL = "https://fsubmqjevlfpcirgsbhi.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZzdWJtcWpldmxmcGNpcmdzYmhpIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDkzNTgyNzgsImV4cCI6MjA2NDkzNDI3OH0.Hfo1kCUCVMvr2ffhLJ3rp7qLMchWYdmBkzOYcorQVQE"

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# Folder containing worksheets
WORKSHEETS_FOLDER = "/Users/safmy/Desktop/Code_and_Scripts/REPOS/train-task-tracker/worktosheets"

# Car type sheets to process (maps to sheet names in Excel)
# Include standard names and De-Icer variants
CAR_SHEETS = [
    "DM 3 Car",
    "Trailer 3 Car",
    "UNDM 3 Car",
    "UNDM 4 Car",
    "Special Trailer 4 Car",
    "Trailer 4 Car",
    "DM 4 Car",
    # De-Icer variants (same car types, different sheet names)
    "De-Icer DM",
    "De-Icer Trailer",
    "DE-Icer UNDM",
]

# Map sheet names to standard car type names for database
SHEET_TO_CAR_TYPE = {
    "DM 3 Car": "DM 3 CAR",
    "Trailer 3 Car": "Trailer 3 Car",
    "UNDM 3 Car": "UNDM 3 CAR",
    "UNDM 4 Car": "UNDM 4 Car",
    "Special Trailer 4 Car": "Special Trailer 4 Car",
    "Trailer 4 Car": "Trailer 4 Car",
    "DM 4 Car": "DM 4 Car",
    # De-Icer mappings to standard names
    "De-Icer DM": "DM 3 CAR",
    "De-Icer Trailer": "Trailer 3 Car",
    "DE-Icer UNDM": "UNDM 3 CAR",
}

# Map sheet names to categories
SHEET_TO_CATEGORY = {
    "DM 3 Car": "3 CAR",
    "Trailer 3 Car": "3 CAR",
    "UNDM 3 Car": "3 CAR",
    "UNDM 4 Car": "4 CAR",
    "Special Trailer 4 Car": "4 CAR",
    "Trailer 4 Car": "4 CAR",
    "DM 4 Car": "4 CAR",
    # De-Icer variants
    "De-Icer DM": "3 CAR",
    "De-Icer Trailer": "3 CAR",
    "DE-Icer UNDM": "3 CAR",
}

# Team mapping - assign initials to teams
# Based on observed patterns, distribute initials across 4 teams
INITIAL_TO_TEAM = {
    # Team A - general group
    "AS": "Team A", "JT": "Team A", "CB": "Team A", "JD": "Team A",
    "KM": "Team A", "CP": "Team A", "KA": "Team A", "TFOS": "Team A",
    # Team B
    "LN": "Team B", "NA": "Team B", "PS": "Team B", "AOO": "Team B",
    "JN": "Team B", "DK": "Team B", "DH": "Team B", "JL": "Team B",
    # Team C
    "SC": "Team C", "MA": "Team C", "CC": "Team C", "OM": "Team C",
    "AL": "Team C", "VN": "Team C", "RN": "Team C", "LVN": "Team C",
    # Team D / Night Shift
    "SA": "Night Shift", "MR": "Night Shift", "AR": "Night Shift", "DB": "Night Shift",
    "GT": "Night Shift", "UQ": "Night Shift", "BP": "Night Shift", "RB": "Night Shift",
}

# Train to unit mapping (fallback for when filename parsing fails)
TRAIN_UNITS = {
    "T01": ["96067", "96122"], "T02": ["96051", "96062"], "T03": ["96099", "96104"],
    "T04": ["96075", "96118"], "T05": ["96113", "96018"], "T06": ["96017", "96078"],
    "T07": ["96011", "96040"], "T08": ["96097", "96086"], "T09": ["96043", "96022"],
    "T10": ["96109", "96110"], "T11": ["96019", "96066"], "T12": ["96101", "96058"],
    "T13": ["96039", "96034"], "T14": ["96055", "96044"], "T15": ["96007", "96096"],
    "T16": ["96057", "96090"], "T17": ["96031", "96032"], "T18": ["96084", "96083"],
    "T19": ["96015", "96082"], "T20": ["96070", "96061"], "T21": ["96063", "96020"],
    "T22": ["96013", "96076"], "T23": ["96103", "96100"], "T24": ["96085", "96106"],
    "T25": ["96081", "96012"], "T26": ["96005", "96008"], "T27": ["96093", "96054"],
    "T28": ["96038", "96003"], "T29": ["96053", "96072"], "T30": ["96041", "96006"],
    "T31": ["96047", "96048"], "T32": ["96123", "96004"], "T33": ["96021", "96094"],
    "T34": ["96025", "96026"], "T35": ["96030", "96029"], "T36": ["96037", "96064"],
    "T37": ["96102", "96089"], "T38": ["96027", "96060"], "T39": ["96024", "96079"],
    "T40": ["96087", "96098"], "T41": ["96119", "96120"], "T42": ["96071", "96116"],
    "T43": ["96073", "96074"], "T44": ["96115", "96124"], "T45": ["96068", "96069"],
    "T46": ["96016", "96117"], "T47": ["96052", "96077"], "T48": ["96105", "96114"],
    "T49": ["96050", "96125"], "T50": ["96010", "96009"], "T51": ["96042", "96107"],
    "T52": ["96028", "96059"], "T53": ["96046", "96045"], "T54": ["96092", "96023"],
    "T55": ["96002", "96001"], "T56": ["96111", "96112"], "T57": ["96014", "96121"],
    "T58": ["96108", "96095"], "T59": ["96088", "96065"], "T60": ["96080", "96091"],
    "T61": ["96036", "96035"], "T62": ["96056", "96033"],
}


def get_train_info_from_filename(filename):
    """Extract train number and units from filename"""
    # Pattern: "WorktosheetsV3 T32 - (Units 96123 & 96004).xlsm"
    # or "WorktosheetsV3.1 T1 - 067&122.xlsm"
    # or "New Work to sheets T2 051 062.xlsm"

    train_match = re.search(r'T(\d+)', filename)
    if train_match:
        train_num = int(train_match.group(1))
        train_id = f"T{train_num:02d}"
    else:
        return None, None, None

    # Try to extract units - multiple patterns
    unit1, unit2 = None, None

    # Pattern 1: Full 5-digit units with & separator - "96123 & 96004"
    units_match = re.search(r'(\d{5})\s*[&]\s*(\d{5})', filename)
    if units_match:
        unit1, unit2 = units_match.group(1), units_match.group(2)

    if not unit1:
        # Pattern 2: Short 3-digit units with & separator - "067&122"
        short_match = re.search(r'(\d{3})\s*[&]\s*(\d{3})', filename)
        if short_match:
            unit1 = f"96{short_match.group(1)}"
            unit2 = f"96{short_match.group(2)}"

    if not unit1:
        # Pattern 3: Short units separated by space - "051 062"
        space_match = re.search(r'T\d+[^0-9]+(\d{3})\s+(\d{3})', filename)
        if space_match:
            unit1 = f"96{space_match.group(1)}"
            unit2 = f"96{space_match.group(2)}"

    if not unit1:
        # Pattern 4: Full units separated by space inside parens - "(96051 96062)"
        paren_match = re.search(r'\((\d{5})\s+(\d{5})\)', filename)
        if paren_match:
            unit1, unit2 = paren_match.group(1), paren_match.group(2)

    if not unit1:
        # Fallback: Use known train-to-unit mapping
        if train_id in TRAIN_UNITS:
            units = TRAIN_UNITS[train_id]
            unit1, unit2 = units[0], units[1]

    return train_id, unit1, unit2


def get_unit_id(unit_number):
    """Get unit ID from Supabase"""
    result = supabase.table('train_units').select('id').eq('unit_number', unit_number).execute()
    if result.data:
        return result.data[0]['id']
    return None


def get_car_type_id(car_type_name):
    """Get car type ID from Supabase"""
    # Normalize name for matching
    normalized = car_type_name.upper().replace(' ', ' ')
    result = supabase.table('car_types').select('id, name').execute()
    for ct in result.data:
        if ct['name'].upper() == normalized or ct['name'].upper().replace(' ', '') == normalized.replace(' ', ''):
            return ct['id']
    return None


def parse_worksheet(file_path):
    """Parse a single worksheet and extract all tasks"""
    filename = os.path.basename(file_path)
    train_id, unit1, unit2 = get_train_info_from_filename(filename)

    if not train_id:
        print(f"  Could not parse train info from: {filename}")
        return []

    print(f"\n  Parsing {train_id} (Units: {unit1}, {unit2})")

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"  Error loading {filename}: {e}")
        return []

    all_tasks = []

    for sheet_name in CAR_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        category = SHEET_TO_CATEGORY.get(sheet_name, "UNKNOWN")
        # Get the standard car type name for database lookup
        car_type_name = SHEET_TO_CAR_TYPE.get(sheet_name, sheet_name)

        # Determine which unit this sheet belongs to (3 CAR or 4 CAR)
        # De-Icer variants are 3 CAR
        is_3_car = "3 Car" in sheet_name or "De-Icer" in sheet_name or "De-icer" in sheet_name
        if is_3_car:
            unit_number = unit1 if unit1 else unit2
        else:  # 4 Car
            unit_number = unit2 if unit2 else unit1

        # NOTE: Removed row 1 unit extraction as Excel sheets often have
        # incorrect unit numbers (e.g., 96411, 96513 which don't exist)
        # Using the units from filename/train mapping instead

        # Parse tasks starting from row 3
        task_count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            # Skip empty rows
            if not row or not any(row[:5]):
                continue

            task_number = str(row[0]) if row[0] else ""
            phase = str(row[1]) if row[1] else ""
            task_name = str(row[2]) if row[2] else ""
            description = str(row[3]) if row[3] else ""
            completed = str(row[4]).lower() if row[4] else ""
            in_progress = str(row[5]).lower() if row[5] else ""
            completed_by = str(row[6]) if row[6] else ""
            date_val = row[7]
            overhaul_iroc = str(row[8]) if len(row) > 8 and row[8] else ""
            position = str(row[9]) if len(row) > 9 and row[9] else ""
            scope_delayed = str(row[10]) if len(row) > 10 and row[10] else ""
            wi_reference = str(row[11]) if len(row) > 11 and row[11] else ""
            num_people = row[13] if len(row) > 13 and row[13] else 1
            total_hours_val = row[14] if len(row) > 14 and row[14] else None

            # Parse total hours (stored as datetime.time like 00:45:00 = 45 minutes)
            total_minutes = 0
            if total_hours_val:
                if isinstance(total_hours_val, dt_time):
                    total_minutes = total_hours_val.hour * 60 + total_hours_val.minute
                elif isinstance(total_hours_val, (int, float)):
                    total_minutes = int(total_hours_val * 60)  # Assume hours as decimal

            # Skip header row or empty task names
            if not task_name or task_name == "Task":
                continue

            # Determine status
            if completed == "yes":
                status = "completed"
            elif in_progress == "yes":
                status = "in_progress"
            else:
                status = "not_started"

            # Parse date
            completed_date = None
            if date_val:
                if isinstance(date_val, datetime):
                    completed_date = date_val.isoformat()
                elif isinstance(date_val, str):
                    try:
                        completed_date = datetime.strptime(date_val, "%Y-%m-%d %H:%M:%S").isoformat()
                    except:
                        pass

            # Extract team from initials
            team_name = None
            if completed_by:
                for initial in completed_by.replace(',', ' ').replace('/', ' ').split():
                    initial = initial.strip().upper()
                    if initial in INITIAL_TO_TEAM:
                        team_name = INITIAL_TO_TEAM[initial]
                        break

            task_data = {
                "train_id": train_id,
                "unit_number": unit_number,
                "car_type": car_type_name,  # Use standardized car type name
                "category": category,
                "task_number": task_number,
                "phase": phase,
                "task_name": task_name[:255] if task_name else "",
                "description": description[:500] if description else "",
                "status": status,
                "completed_by": completed_by[:100] if completed_by else "",
                "completed_date": completed_date,
                "overhaul_iroc": overhaul_iroc[:50] if overhaul_iroc else "",
                "position": position[:50] if position else "",
                "scope_delayed": scope_delayed.upper() == "Y" if scope_delayed else False,
                "wi_reference": wi_reference[:255] if wi_reference else "",
                "team_name": team_name,  # Add team based on initials
                "num_people": int(num_people) if num_people else 1,
                "total_minutes": total_minutes,  # Task duration in minutes
            }

            all_tasks.append(task_data)
            task_count += 1

        if task_count > 0:
            print(f"    {sheet_name}: {task_count} tasks")

    wb.close()
    return all_tasks


def upload_to_supabase(all_tasks):
    """Upload all tasks to Supabase"""
    print(f"\n{'='*60}")
    print(f"UPLOADING {len(all_tasks)} TASKS TO SUPABASE")
    print(f"{'='*60}")

    # Get unit IDs mapping
    units_result = supabase.table('train_units').select('id, unit_number').execute()
    unit_id_map = {u['unit_number']: u['id'] for u in units_result.data}

    # Get car type IDs mapping
    car_types_result = supabase.table('car_types').select('id, name').execute()
    car_type_id_map = {}
    for ct in car_types_result.data:
        car_type_id_map[ct['name'].upper()] = ct['id']
        # Also add normalized versions
        normalized = ct['name'].upper().replace(' ', '')
        car_type_id_map[normalized] = ct['id']

    # Prepare tasks for upload
    tasks_to_upload = []
    skipped = 0

    for task in all_tasks:
        unit_id = unit_id_map.get(task['unit_number'])
        if not unit_id:
            skipped += 1
            continue

        # Match car type
        car_type_key = task['car_type'].upper()
        car_type_id = car_type_id_map.get(car_type_key)
        if not car_type_id:
            car_type_key = car_type_key.replace(' ', '')
            car_type_id = car_type_id_map.get(car_type_key)

        upload_task = {
            "unit_id": unit_id,
            "car_type_id": car_type_id,
            "task_number": task['task_number'],
            "phase": task['phase'],
            "task_name": task['task_name'],
            "description": task['description'],
            "status": task['status'],
            "completed_by": task['completed_by'] or None,
            "completed_date": task['completed_date'],
            "overhaul_iroc": task['overhaul_iroc'] or None,
            "position": task['position'] or None,
            "scope_delayed": task['scope_delayed'],
            "wi_reference": task['wi_reference'] or None,
            "total_minutes": task.get('total_minutes', 0),
            "num_people": task.get('num_people', 1),
        }
        tasks_to_upload.append(upload_task)

    print(f"Tasks ready for upload: {len(tasks_to_upload)}")
    print(f"Skipped (no unit match): {skipped}")

    # Upload in batches
    batch_size = 500
    uploaded = 0
    errors = 0

    for i in range(0, len(tasks_to_upload), batch_size):
        batch = tasks_to_upload[i:i+batch_size]
        try:
            result = supabase.table('tasks').upsert(batch).execute()
            uploaded += len(batch)
            print(f"  Uploaded batch {i//batch_size + 1}: {len(batch)} tasks")
        except Exception as e:
            errors += len(batch)
            print(f"  Error uploading batch: {e}")

    print(f"\nUpload complete: {uploaded} success, {errors} errors")
    return uploaded, errors


def main():
    print("="*60)
    print("PARSING WORKTOSHEETS AND UPLOADING TO SUPABASE")
    print("="*60)

    # Get all Excel files
    files = [f for f in os.listdir(WORKSHEETS_FOLDER) if f.endswith('.xlsm')]
    files.sort()

    print(f"\nFound {len(files)} worksheet files")

    # Parse all files
    all_tasks = []
    for idx, filename in enumerate(files, 1):
        file_path = os.path.join(WORKSHEETS_FOLDER, filename)
        print(f"\n[{idx}/{len(files)}] {filename}")
        tasks = parse_worksheet(file_path)
        all_tasks.extend(tasks)

    print(f"\n{'='*60}")
    print(f"TOTAL TASKS EXTRACTED: {len(all_tasks)}")
    print(f"{'='*60}")

    # Summary by status
    status_counts = {}
    for task in all_tasks:
        status = task['status']
        status_counts[status] = status_counts.get(status, 0) + 1

    print("\nStatus breakdown:")
    for status, count in sorted(status_counts.items()):
        print(f"  {status}: {count}")

    # Upload to Supabase
    uploaded, errors = upload_to_supabase(all_tasks)

    return all_tasks


if __name__ == "__main__":
    main()
