#!/usr/bin/env python3
"""
Migrate data from tasks table to cars + task_completions tables
"""

from supabase import create_client
from datetime import datetime
import openpyxl
import os

# Supabase configuration
SUPABASE_URL = "https://fsubmqjevlfpcirgsbhi.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZzdWJtcWpldmxmcGNpcmdzYmhpIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDkzNTgyNzgsImV4cCI6MjA2NDkzNDI3OH0.Hfo1kCUCVMvr2ffhLJ3rp7qLMchWYdmBkzOYcorQVQE"

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

def load_phase_mapping():
    """Load task -> phase mapping from Master Data sheet"""
    master_file = "Work2Sheets Masters.xlsx"
    if not os.path.exists(master_file):
        print(f"  Warning: {master_file} not found, phases will be empty")
        return {}

    print(f"  Loading phase mapping from {master_file}...")
    wb = openpyxl.load_workbook(master_file, data_only=True)

    if 'Master Data' not in wb.sheetnames:
        print("  Warning: 'Master Data' sheet not found")
        return {}

    sheet = wb['Master Data']
    task_phases = {}

    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        phase, task = row
        if phase and task:
            # Normalize phase names (capitalize properly, handle variations)
            phase_str = str(phase).strip()
            # Normalize catchback variations
            if phase_str.lower() in ['catch back', 'catchback']:
                phase_str = 'Catchback'
            task_phases[str(task).strip().upper()] = phase_str

    print(f"  Loaded {len(task_phases)} task->phase mappings")

    # Print phase distribution
    phase_counts = {}
    for phase in task_phases.values():
        phase_counts[phase] = phase_counts.get(phase, 0) + 1
    print(f"  Phase distribution: {dict(sorted(phase_counts.items()))}")

    return task_phases

# Team mapping - assign initials to teams
INITIAL_TO_TEAM = {
    # Team A - general group
    "AS": "Team A", "JT": "Team A", "CB": "Team A", "JD": "Team A",
    "KM": "Team A", "CP": "Team A", "KA": "Team A",
    # Team B
    "LN": "Team B", "NA": "Team B", "PS": "Team B", "AOO": "Team B",
    "JN": "Team B", "DK": "Team B", "DH": "Team B", "JL": "Team B",
    # Team C
    "SC": "Team C", "MA": "Team C", "CC": "Team C", "OM": "Team C",
    "AL": "Team C", "VN": "Team C", "RN": "Team C", "LVN": "Team C",
    # Night Shift (Team D)
    "SA": "Night Shift", "MR": "Night Shift", "AR": "Night Shift", "DB": "Night Shift",
    "GT": "Night Shift", "UQ": "Night Shift", "BP": "Night Shift", "RB": "Night Shift",
    # TFOS - Unknown person, separate category until identified
    "TFOS": "TFOS",
}


def migrate():
    print("=" * 60)
    print("MIGRATING TASKS TO CARS + TASK_COMPLETIONS")
    print("=" * 60)

    # Load phase mapping from Master Data
    print("\nLoading phase mapping...")
    task_phases = load_phase_mapping()

    # Get team IDs
    print("\nFetching teams...")
    teams_result = supabase.table('teams').select('id, name').execute()
    team_id_map = {t['name']: t['id'] for t in teams_result.data}
    print(f"  Teams: {list(team_id_map.keys())}")

    # Get all tasks
    print("\nFetching tasks...")
    all_tasks = []
    offset = 0
    batch_size = 1000

    while True:
        result = supabase.table('tasks').select('*').range(offset, offset + batch_size - 1).execute()
        if not result.data:
            break
        all_tasks.extend(result.data)
        offset += batch_size
        print(f"  Fetched {len(all_tasks)} tasks...")

    print(f"\nTotal tasks to migrate: {len(all_tasks)}")

    # Clear existing cars and task_completions
    print("\nClearing existing data...")
    supabase.table('task_completions').delete().neq('id', '00000000-0000-0000-0000-000000000000').execute()
    supabase.table('cars').delete().neq('id', '00000000-0000-0000-0000-000000000000').execute()
    print("  Cleared existing cars and task_completions")

    # Group tasks by unit_id + car_type_id
    car_tasks = {}
    for task in all_tasks:
        key = (task['unit_id'], task['car_type_id'])
        if key not in car_tasks:
            car_tasks[key] = []
        car_tasks[key].append(task)

    print(f"\nUnique car combinations: {len(car_tasks)}")

    # Create cars and task_completions
    cars_created = 0
    completions_created = 0

    for (unit_id, car_type_id), tasks in car_tasks.items():
        if not unit_id or not car_type_id:
            continue

        # Generate car number from task number (e.g., first task's task_number or sequential)
        car_number = tasks[0].get('task_number', '') or str(cars_created + 1)

        # Create car
        try:
            car_result = supabase.table('cars').insert({
                'unit_id': unit_id,
                'car_type_id': car_type_id,
                'car_number': car_number[:50]
            }).execute()

            if not car_result.data:
                continue

            car_id = car_result.data[0]['id']
            cars_created += 1

            # Create task_completions in batches
            completions = []
            for idx, task in enumerate(tasks):
                # Map status
                status = task.get('status', 'pending')
                if status == 'not_started':
                    status = 'pending'

                # Parse completed_by
                completed_by = []
                if task.get('completed_by'):
                    completed_by = [s.strip() for s in str(task['completed_by']).split(',') if s.strip()]

                # Get team_id from initials in completed_by
                team_id = None
                if completed_by:
                    for initial in completed_by:
                        initial_upper = initial.upper()
                        if initial_upper in INITIAL_TO_TEAM:
                            team_name = INITIAL_TO_TEAM[initial_upper]
                            if team_name in team_id_map:
                                team_id = team_id_map[team_name]
                                break

                # Look up phase from task name
                task_name_upper = task.get('task_name', '').strip().upper()
                phase = task_phases.get(task_name_upper, None)

                completion = {
                    'car_id': car_id,
                    'task_name': task.get('task_name', '')[:255],
                    'description': task.get('description', '')[:500] if task.get('description') else None,
                    'status': status,
                    'completed_by': completed_by if completed_by else None,
                    'completed_at': task.get('completed_date'),
                    'sort_order': idx + 1,
                    'team_id': team_id,
                    'total_minutes': task.get('total_minutes', 0),
                    'num_people': task.get('num_people', 1),
                    'phase': phase,
                }
                completions.append(completion)

            # Insert completions in batches
            batch_size = 100
            for i in range(0, len(completions), batch_size):
                batch = completions[i:i + batch_size]
                try:
                    supabase.table('task_completions').insert(batch).execute()
                    completions_created += len(batch)
                except Exception as e:
                    print(f"  Error inserting completions batch: {e}")

            if cars_created % 20 == 0:
                print(f"  Progress: {cars_created} cars, {completions_created} completions")

        except Exception as e:
            print(f"  Error creating car for unit {unit_id}: {e}")
            continue

    print("\n" + "=" * 60)
    print(f"MIGRATION COMPLETE")
    print(f"  Cars created: {cars_created}")
    print(f"  Task completions created: {completions_created}")
    print("=" * 60)

    # Verify
    cars_count = supabase.table('cars').select('id', count='exact').execute()
    completions_count = supabase.table('task_completions').select('id', count='exact').execute()
    print(f"\nVerification:")
    print(f"  Cars in database: {cars_count.count}")
    print(f"  Task completions in database: {completions_count.count}")

    # Check phase distribution in database
    print("\nPhase distribution in database:")
    phases_result = supabase.table('task_completions').select('phase').execute()
    phase_counts = {}
    for row in phases_result.data:
        phase = row.get('phase') or 'No Phase'
        phase_counts[phase] = phase_counts.get(phase, 0) + 1
    for phase, count in sorted(phase_counts.items()):
        print(f"  {phase}: {count}")


if __name__ == "__main__":
    migrate()
