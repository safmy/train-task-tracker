#!/usr/bin/env python3
"""
Fast update of total_minutes in task_completions from Work2Sheets Masters.xlsx
Uses batch updates by task_name instead of individual record updates.
"""

import openpyxl
from datetime import time, datetime
from supabase import create_client
from collections import defaultdict
import sys

# Force unbuffered output
sys.stdout.reconfigure(line_buffering=True)

# Supabase configuration
SUPABASE_URL = "https://fsubmqjevlfpcirgsbhi.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZzdWJtcWpldmxmcGNpcmdzYmhpIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDkzNTgyNzgsImV4cCI6MjA2NDkzNDI3OH0.Hfo1kCUCVMvr2ffhLJ3rp7qLMchWYdmBkzOYcorQVQE"

def parse_timing(value):
    """Parse timing value from Excel. Format is HH:MM:SS."""
    if value is None:
        return None

    if isinstance(value, time):
        total_mins = value.hour * 60 + value.minute
        if total_mins == 0 and value.second > 0:
            total_mins = value.second
        return total_mins

    if isinstance(value, datetime):
        total_mins = value.hour * 60 + value.minute
        if total_mins == 0 and value.second > 0:
            total_mins = value.second
        return total_mins if total_mins > 0 else None

    if isinstance(value, str):
        try:
            parts = value.split(':')
            if len(parts) == 3:
                h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
                total_mins = h * 60 + m
                if total_mins == 0 and s > 0:
                    total_mins = s
                return total_mins
            elif len(parts) == 2:
                h, m = int(parts[0]), int(parts[1])
                return h * 60 + m
        except:
            pass

    if isinstance(value, (int, float)):
        if value < 1:
            return round(value * 24 * 60)
        else:
            return round(value)

    return None


def main():
    print("=" * 70)
    print("FAST UPDATE TASK MINUTES FROM EXCEL")
    print("=" * 70)
    print(flush=True)

    # Step 1: Load timing data from Excel
    print("1. Loading timing data from Work2Sheets Masters.xlsx...")
    print(flush=True)

    wb = openpyxl.load_workbook(
        '/Users/safmy/Desktop/Code_and_Scripts/REPOS/train-task-tracker/Work2Sheets Masters.xlsx',
        data_only=True
    )
    sheet = wb['Master Data']

    # Extract unique task -> minutes mapping
    task_timings = {}

    for row_num in range(2, sheet.max_row + 1):
        task = sheet.cell(row=row_num, column=2).value  # Column B - Task
        timing = sheet.cell(row=row_num, column=14).value  # Column N - Total Hours

        if not task:
            continue

        task_upper = str(task).strip().upper()
        minutes = parse_timing(timing)

        if minutes is not None and task_upper not in task_timings:
            task_timings[task_upper] = minutes

    print(f"   Found {len(task_timings)} unique tasks with timing data")
    print(flush=True)

    # Step 2: Connect to Supabase
    print("\n2. Connecting to Supabase...")
    print(flush=True)
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Step 3: Update by task_name in batches
    print("\n3. Updating task_completions by task_name...")
    print(f"   Processing {len(task_timings)} unique task names...")
    print(flush=True)

    updated_count = 0
    error_count = 0

    task_list = list(task_timings.items())

    for i, (task_name, minutes) in enumerate(task_list):
        try:
            # Update all records with this task_name
            result = supabase.table('task_completions').update({
                'total_minutes': minutes
            }).ilike('task_name', task_name).execute()

            # Count affected rows (Supabase returns updated records)
            if result.data:
                updated_count += len(result.data)

            if (i + 1) % 50 == 0:
                print(f"   Processed {i + 1}/{len(task_list)} task types, updated {updated_count} records...")
                print(flush=True)

        except Exception as e:
            error_count += 1
            if error_count <= 5:
                print(f"   Error updating '{task_name[:40]}': {e}")
                print(flush=True)

    print(f"\n   Done! Processed: {len(task_list)} task types")
    print(f"   Records updated: {updated_count}")
    print(f"   Errors: {error_count}")
    print(flush=True)

    # Step 4: Verify
    print("\n4. Verifying...")
    print(flush=True)

    verify = supabase.table('task_completions').select(
        'id', count='exact'
    ).gt('total_minutes', 0).execute()
    print(f"   Tasks with total_minutes > 0: {verify.count}")
    print(flush=True)


if __name__ == "__main__":
    main()
