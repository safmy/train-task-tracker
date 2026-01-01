#!/usr/bin/env python3
"""
Update total_minutes in task_completions from Work2Sheets Masters.xlsx
"""

import openpyxl
from datetime import time, datetime
from supabase import create_client
from collections import defaultdict

# Supabase configuration
SUPABASE_URL = "https://fsubmqjevlfpcirgsbhi.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZzdWJtcWpldmxmcGNpcmdzYmhpIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDkzNTgyNzgsImV4cCI6MjA2NDkzNDI3OH0.Hfo1kCUCVMvr2ffhLJ3rp7qLMchWYdmBkzOYcorQVQE"

def parse_timing(value):
    """
    Parse timing value from Excel.
    Format is HH:MM:SS where we want total minutes.

    For datetime values (Excel corruption), we try to extract the time part.
    For values like 1900-02-14 where 45 was entered, we can't recover the original.
    """
    if value is None:
        return None

    if isinstance(value, time):
        # HH:MM:SS -> convert to minutes
        # Note: seconds component seems to sometimes contain minutes
        # e.g., 00:00:45 might mean 45 minutes, not 45 seconds
        total_mins = value.hour * 60 + value.minute
        if total_mins == 0 and value.second > 0:
            # If hours and minutes are 0 but seconds isn't,
            # the seconds might actually be minutes
            total_mins = value.second
        return total_mins

    if isinstance(value, datetime):
        # Excel corruption - extract time part
        total_mins = value.hour * 60 + value.minute
        if total_mins == 0 and value.second > 0:
            total_mins = value.second
        # If still 0 but date is weird, might be encoded in date
        if total_mins == 0 and value.day > 1:
            # Days since 1900-01-01 might represent hours or minutes
            # This is a guess - flag for manual review
            pass
        return total_mins if total_mins > 0 else None

    if isinstance(value, str):
        # Parse HH:MM:SS string
        try:
            parts = value.split(':')
            if len(parts) == 3:
                h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
                total_mins = h * 60 + m
                if total_mins == 0 and s > 0:
                    total_mins = s
                return total_mins
            elif len(parts) == 2:
                h, m = int(parts[0]), int(parts[1])
                return h * 60 + m
        except:
            pass

    if isinstance(value, (int, float)):
        # Numeric - could be fraction of a day or direct minutes
        if value < 1:
            # Fraction of a day
            return round(value * 24 * 60)
        else:
            # Might be direct minutes
            return round(value)

    return None


def main():
    print("=" * 70)
    print("UPDATE TASK MINUTES FROM EXCEL")
    print("=" * 70)

    # Step 1: Load timing data from Excel
    print("\n1. Loading timing data from Work2Sheets Masters.xlsx...")

    wb = openpyxl.load_workbook(
        '/Users/safmy/Desktop/Code_and_Scripts/REPOS/train-task-tracker/Work2Sheets Masters.xlsx',
        data_only=True
    )
    sheet = wb['Master Data']

    # Extract unique task -> minutes mapping
    task_timings = {}
    timing_issues = []

    for row_num in range(2, sheet.max_row + 1):
        task = sheet.cell(row=row_num, column=2).value  # Column B - Task
        timing = sheet.cell(row=row_num, column=14).value  # Column N - Total Hours
        num_people = sheet.cell(row=row_num, column=13).value  # Column M - Number Of People

        if not task:
            continue

        task_upper = str(task).strip().upper()
        minutes = parse_timing(timing)

        if minutes is not None:
            if task_upper not in task_timings:
                task_timings[task_upper] = {
                    'minutes': minutes,
                    'num_people': num_people if isinstance(num_people, (int, float)) else 1,
                    'raw': timing
                }
        else:
            if timing is not None and task_upper not in [t[0] for t in timing_issues]:
                timing_issues.append((task_upper[:50], timing, type(timing).__name__))

    print(f"   Found {len(task_timings)} unique tasks with timing data")

    # Show sample
    print("\n   Sample task timings:")
    for i, (task, data) in enumerate(list(task_timings.items())[:10]):
        print(f"   - {task[:55]:<55} = {data['minutes']:>4} mins (people: {data['num_people']})")

    # Show timing distribution
    print("\n   Timing distribution:")
    dist = defaultdict(int)
    for task, data in task_timings.items():
        mins = data['minutes']
        if mins == 0:
            dist['0 mins'] += 1
        elif mins <= 15:
            dist['1-15 mins'] += 1
        elif mins <= 30:
            dist['16-30 mins'] += 1
        elif mins <= 60:
            dist['31-60 mins'] += 1
        elif mins <= 120:
            dist['1-2 hours'] += 1
        else:
            dist['>2 hours'] += 1

    for bucket in ['0 mins', '1-15 mins', '16-30 mins', '31-60 mins', '1-2 hours', '>2 hours']:
        if bucket in dist:
            print(f"   {bucket}: {dist[bucket]}")

    if timing_issues:
        print(f"\n   Warning: {len(timing_issues)} tasks had unparseable timing:")
        for task, val, typ in timing_issues[:5]:
            print(f"   - {task}: {val} ({typ})")

    # Step 2: Connect to Supabase
    print("\n2. Connecting to Supabase...")
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Step 3: Get unique task names from database
    print("\n3. Fetching task names from database...")

    all_completions = []
    offset = 0
    while True:
        result = supabase.table('task_completions').select(
            'id, task_name, total_minutes'
        ).range(offset, offset + 999).execute()
        if not result.data:
            break
        all_completions.extend(result.data)
        offset += 1000
        if offset % 10000 == 0:
            print(f"   Fetched {offset} records...")

    print(f"   Total task completions: {len(all_completions)}")

    # Count tasks with/without minutes
    has_minutes = sum(1 for c in all_completions if c.get('total_minutes') and c['total_minutes'] > 0)
    needs_update = sum(1 for c in all_completions if not c.get('total_minutes') or c['total_minutes'] == 0)

    print(f"   Tasks with minutes: {has_minutes}")
    print(f"   Tasks needing update: {needs_update}")

    # Step 4: Match and prepare updates
    print("\n4. Matching database tasks to Excel timings...")

    updates = []
    matched = 0
    unmatched_tasks = set()

    for c in all_completions:
        task_name = c.get('task_name', '').strip().upper()
        current_mins = c.get('total_minutes') or 0

        if task_name in task_timings:
            new_mins = task_timings[task_name]['minutes']
            if new_mins != current_mins:
                updates.append({
                    'id': c['id'],
                    'total_minutes': new_mins
                })
            matched += 1
        else:
            if task_name:
                unmatched_tasks.add(task_name[:60])

    print(f"   Matched: {matched}")
    print(f"   Unmatched: {len(unmatched_tasks)}")
    print(f"   Updates needed: {len(updates)}")

    if unmatched_tasks:
        print(f"\n   Sample unmatched tasks (first 10):")
        for task in list(unmatched_tasks)[:10]:
            print(f"   - {task}")

    if not updates:
        print("\n   No updates needed!")
        return

    # Step 5: Apply updates
    print(f"\n5. Updating {len(updates)} task_completions with timing data...")

    print("   Updating...")
    updated = 0
    errors = 0

    for i, update in enumerate(updates):
        try:
            supabase.table('task_completions').update({
                'total_minutes': update['total_minutes']
            }).eq('id', update['id']).execute()
            updated += 1

            if updated % 1000 == 0:
                print(f"   Updated {updated}/{len(updates)}...")
        except Exception as e:
            errors += 1
            if errors <= 3:
                print(f"   Error: {e}")

    print(f"\n   Done! Updated: {updated}, Errors: {errors}")

    # Verify
    print("\n6. Verifying...")
    verify = supabase.table('task_completions').select(
        'id', count='exact'
    ).gt('total_minutes', 0).execute()
    print(f"   Tasks with total_minutes > 0: {verify.count}")


if __name__ == "__main__":
    main()
