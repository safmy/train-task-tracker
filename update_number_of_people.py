#!/usr/bin/env python3
"""
Update task_completions with number_of_people from Work2Sheets Masters.xlsx
This multiplier is needed for accurate efficiency calculation.
A 7-hour job with 2 people = 14 man-hours of work.
"""

import openpyxl
from datetime import time, datetime
from supabase import create_client
import sys

# Force unbuffered output
sys.stdout.reconfigure(line_buffering=True)

# Supabase configuration
SUPABASE_URL = "https://fsubmqjevlfpcirgsbhi.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZzdWJtcWpldmxmcGNpcmdzYmhpIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDkzNTgyNzgsImV4cCI6MjA2NDkzNDI3OH0.Hfo1kCUCVMvr2ffhLJ3rp7qLMchWYdmBkzOYcorQVQE"


def main():
    print("=" * 70)
    print("UPDATE NUMBER OF PEOPLE FROM EXCEL")
    print("=" * 70)
    print(flush=True)

    # Step 1: Load data from Excel
    print("1. Loading data from Work2Sheets Masters.xlsx...")
    print(flush=True)

    wb = openpyxl.load_workbook(
        '/Users/safmy/Desktop/Code_and_Scripts/REPOS/train-task-tracker/Work2Sheets Masters.xlsx',
        data_only=True
    )
    sheet = wb['Master Data']

    # Extract task -> number_of_people mapping
    task_people = {}

    for row_num in range(2, sheet.max_row + 1):
        task = sheet.cell(row=row_num, column=2).value  # Column B - Task
        people = sheet.cell(row=row_num, column=13).value  # Column M - Number Of People

        if not task:
            continue

        task_upper = str(task).strip().upper()

        # Parse number of people (default to 1)
        num_people = 1
        if people is not None:
            try:
                num_people = int(people)
                if num_people < 1:
                    num_people = 1
            except (ValueError, TypeError):
                num_people = 1

        # Only store if we haven't seen this task or if this has more people
        if task_upper not in task_people:
            task_people[task_upper] = num_people

    print(f"   Found {len(task_people)} unique tasks with people data")

    # Show sample data
    print("\n   Sample data:")
    for i, (task, people) in enumerate(list(task_people.items())[:5]):
        print(f"     {task[:50]}: {people} people")
    print(flush=True)

    # Step 2: Connect to Supabase
    print("\n2. Connecting to Supabase...")
    print(flush=True)
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Step 3: Update by task_name in batches
    print("\n3. Updating task_completions by task_name...")
    print(f"   Processing {len(task_people)} unique task names...")
    print(flush=True)

    updated_count = 0
    error_count = 0

    task_list = list(task_people.items())

    for i, (task_name, num_people) in enumerate(task_list):
        try:
            # Update all records with this task_name
            result = supabase.table('task_completions').update({
                'number_of_people': num_people
            }).ilike('task_name', task_name).execute()

            # Count affected rows
            if result.data:
                updated_count += len(result.data)

            if (i + 1) % 50 == 0:
                print(f"   Processed {i + 1}/{len(task_list)} task types, updated {updated_count} records...")
                print(flush=True)

        except Exception as e:
            error_count += 1
            if error_count <= 5:
                print(f"   Error updating '{task_name[:40]}': {e}")
                print(flush=True)

    print(f"\n   Done! Processed: {len(task_list)} task types")
    print(f"   Records updated: {updated_count}")
    print(f"   Errors: {error_count}")
    print(flush=True)

    # Step 4: Verify
    print("\n4. Verifying...")
    print(flush=True)

    verify = supabase.table('task_completions').select(
        'id', count='exact'
    ).gt('number_of_people', 1).execute()
    print(f"   Tasks with number_of_people > 1: {verify.count}")

    # Show distribution
    for num in [1, 2, 3, 4]:
        count_result = supabase.table('task_completions').select(
            'id', count='exact'
        ).eq('number_of_people', num).execute()
        print(f"   Tasks with {num} people: {count_result.count}")
    print(flush=True)


if __name__ == "__main__":
    main()
