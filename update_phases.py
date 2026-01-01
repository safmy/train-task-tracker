#!/usr/bin/env python3
"""
Update existing task_completions with phase information from Master Data
"""

from supabase import create_client
import openpyxl
import os

# Supabase configuration
SUPABASE_URL = "https://fsubmqjevlfpcirgsbhi.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZzdWJtcWpldmxmcGNpcmdzYmhpIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDkzNTgyNzgsImV4cCI6MjA2NDkzNDI3OH0.Hfo1kCUCVMvr2ffhLJ3rp7qLMchWYdmBkzOYcorQVQE"

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

def load_phase_mapping():
    """Load task -> phase mapping from Master Data sheet"""
    master_file = "Work2Sheets Masters.xlsx"
    if not os.path.exists(master_file):
        print(f"Error: {master_file} not found")
        return {}

    print(f"Loading phase mapping from {master_file}...")
    wb = openpyxl.load_workbook(master_file, data_only=True)

    if 'Master Data' not in wb.sheetnames:
        print("Error: 'Master Data' sheet not found")
        return {}

    sheet = wb['Master Data']
    task_phases = {}

    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        phase, task = row
        if phase and task:
            phase_str = str(phase).strip()
            if phase_str.lower() in ['catch back', 'catchback']:
                phase_str = 'Catchback'
            task_phases[str(task).strip().upper()] = phase_str

    print(f"Loaded {len(task_phases)} task->phase mappings")

    # Print phase distribution
    phase_counts = {}
    for phase in task_phases.values():
        phase_counts[phase] = phase_counts.get(phase, 0) + 1
    print(f"Phase distribution: {dict(sorted(phase_counts.items()))}")

    return task_phases


def update_phases():
    print("=" * 60)
    print("UPDATING TASK_COMPLETIONS WITH PHASES")
    print("=" * 60)

    # Load phase mapping
    task_phases = load_phase_mapping()
    if not task_phases:
        return

    # Get all unique task names from task_completions
    print("\nFetching unique task names...")
    all_tasks = []
    offset = 0
    batch_size = 1000

    while True:
        result = supabase.table('task_completions').select('id, task_name').range(offset, offset + batch_size - 1).execute()
        if not result.data:
            break
        all_tasks.extend(result.data)
        offset += batch_size
        if offset % 10000 == 0:
            print(f"  Fetched {len(all_tasks)} tasks...")

    print(f"Total task_completions: {len(all_tasks)}")

    # Group by task_name for efficiency
    tasks_by_name = {}
    for task in all_tasks:
        name = task['task_name'].strip().upper()
        if name not in tasks_by_name:
            tasks_by_name[name] = []
        tasks_by_name[name].append(task['id'])

    print(f"Unique task names: {len(tasks_by_name)}")

    # Update phases in batches
    updated = 0
    matched = 0
    unmatched_names = set()

    for task_name, task_ids in tasks_by_name.items():
        phase = task_phases.get(task_name)

        if phase:
            matched += len(task_ids)
            # Update in batches of 100 IDs
            for i in range(0, len(task_ids), 100):
                batch_ids = task_ids[i:i+100]
                try:
                    supabase.table('task_completions').update({'phase': phase}).in_('id', batch_ids).execute()
                    updated += len(batch_ids)
                except Exception as e:
                    print(f"  Error updating batch: {e}")
        else:
            unmatched_names.add(task_name[:50])

        if updated > 0 and updated % 5000 == 0:
            print(f"  Progress: {updated} tasks updated...")

    print("\n" + "=" * 60)
    print("UPDATE COMPLETE")
    print(f"  Tasks matched with phases: {matched}")
    print(f"  Tasks updated: {updated}")
    print(f"  Unmatched task names: {len(unmatched_names)}")
    print("=" * 60)

    if unmatched_names:
        print("\nSample unmatched task names (first 10):")
        for name in list(unmatched_names)[:10]:
            print(f"  - {name}")

    # Verify phase distribution
    print("\nVerifying phase distribution in database...")
    phases_result = supabase.table('task_completions').select('phase').limit(1000).execute()
    phase_counts = {}
    for row in phases_result.data:
        phase = row.get('phase') or 'No Phase'
        phase_counts[phase] = phase_counts.get(phase, 0) + 1
    print("Sample phase distribution (first 1000 rows):")
    for phase, count in sorted(phase_counts.items()):
        print(f"  {phase}: {count}")


if __name__ == "__main__":
    update_phases()
